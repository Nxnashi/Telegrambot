import io
import logging
import os

from flask import request, jsonify, send_from_directory, send_file

from config import TOKEN, OPERATOR_IDS, ADMIN_IDS
from webapp import validate_init_data
from database import get_dashboard_requests, get_all_operators_stats, get_all_requests, get_events
from handlers.operator_handlers import take_request, complete_request, postpone_request, resume_request, cancel_request, restore_request

logger = logging.getLogger(__name__)

DASHBOARD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "dashboard")


def _check_operator(init_data):
    """Проверяет подпись и что пользователь — оператор. Возвращает user dict или None."""
    user = validate_init_data(init_data, TOKEN)
    if not user:
        return None
    if user.get("id") not in OPERATOR_IDS:
        return None
    return user


def register_dashboard(app, bot):

    @app.route("/dashboard/")
    @app.route("/dashboard")
    def dashboard_index():
        return send_from_directory(DASHBOARD_DIR, "index.html")

    @app.route("/dashboard/<path:filename>")
    def dashboard_static(filename):
        return send_from_directory(DASHBOARD_DIR, filename)

    # =========================
    # КТО Я (оператор / админ)
    # =========================
    @app.route("/api/dashboard/me")
    def api_me():
        init_data = request.args.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False), 403

        return jsonify(
            ok=True,
            name=user.get("first_name", "Оператор"),
            is_admin=user.get("id") in ADMIN_IDS
        )

    # =========================
    # СПИСОК ЗАЯВОК ДЛЯ ДОСКИ
    # =========================
    @app.route("/api/dashboard/requests")
    def api_requests():
        init_data = request.args.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False), 403

        rows = get_dashboard_requests()
        items = [
            {
                "id": r[0],
                "restaurant": r[1],
                "status": r[2],
                "operator_name": r[3],
                "description": r[4],
                "rating": r[5],
                "name": r[6],
                "phone": r[7],
                "reason": r[8],
            }
            for r in rows
        ]
        return jsonify(ok=True, items=items)

    # =========================
    # ЖУРНАЛ СОБЫТИЙ (видно всем операторам)
    # =========================
    @app.route("/api/dashboard/events")
    def api_events():
        init_data = request.args.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False), 403

        request_id = request.args.get("request_id")
        rows = get_events(request_id=request_id)
        items = [
            {
                "id": r[0],
                "request_id": r[1],
                "event_type": r[2],
                "actor_name": r[3],
                "details": r[4],
                "created_at": r[5],
            }
            for r in rows
        ]
        return jsonify(ok=True, items=items)

    # =========================
    # ОТЛОЖИТЬ ЗАЯВКУ
    # =========================
    @app.route("/api/dashboard/postpone", methods=["POST"])
    def api_postpone():
        init_data = request.form.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False, error="forbidden"), 403

        request_id = request.form.get("request_id")
        reason = request.form.get("reason", "")
        operator_name = user.get("first_name", "Оператор")
        ok, message = postpone_request(bot, request_id, user["id"], operator_name, reason)
        return jsonify(ok=ok, message=message)

    # =========================
    # ВОЗОБНОВИТЬ ЗАЯВКУ
    # =========================
    @app.route("/api/dashboard/resume", methods=["POST"])
    def api_resume():
        init_data = request.form.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False, error="forbidden"), 403

        request_id = request.form.get("request_id")
        operator_name = user.get("first_name", "Оператор")
        ok, message = resume_request(bot, request_id, user["id"], operator_name)
        return jsonify(ok=ok, message=message)

    # =========================
    # ОТМЕНИТЬ ЗАЯВКУ
    # =========================
    @app.route("/api/dashboard/cancel", methods=["POST"])
    def api_cancel():
        init_data = request.form.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False, error="forbidden"), 403

        request_id = request.form.get("request_id")
        reason = request.form.get("reason", "")
        operator_name = user.get("first_name", "Оператор")
        ok, message = cancel_request(bot, request_id, user["id"], operator_name, reason)
        return jsonify(ok=ok, message=message)

    # =========================
    # ВЕРНУТЬ ОТМЕНЁННУЮ ЗАЯВКУ
    # =========================
    @app.route("/api/dashboard/restore", methods=["POST"])
    def api_restore():
        init_data = request.form.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False, error="forbidden"), 403

        request_id = request.form.get("request_id")
        operator_name = user.get("first_name", "Оператор")
        ok, message = restore_request(bot, request_id, user["id"], operator_name)
        return jsonify(ok=ok, message=message)

    # =========================
    # ВЗЯТЬ ЗАЯВКУ (перетащили в "В процессе")
    # =========================
    @app.route("/api/dashboard/take", methods=["POST"])
    def api_take():
        init_data = request.form.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False, error="forbidden"), 403

        request_id = request.form.get("request_id")
        operator_name = user.get("first_name", "Оператор")
        ok, message = take_request(bot, request_id, user["id"], operator_name)
        return jsonify(ok=ok, message=message)

    # =========================
    # ЗАВЕРШИТЬ ЗАЯВКУ (перетащили в "Выполнено")
    # =========================
    @app.route("/api/dashboard/complete", methods=["POST"])
    def api_complete():
        init_data = request.form.get("initData", "")
        user = _check_operator(init_data)
        if not user:
            return jsonify(ok=False, error="forbidden"), 403

        request_id = request.form.get("request_id")
        operator_name = user.get("first_name", "Оператор")
        ok, message = complete_request(bot, request_id, user["id"], operator_name)
        return jsonify(ok=ok, message=message)

    # =========================
    # СТАТИСТИКА (только админ)
    # =========================
    @app.route("/api/dashboard/stats")
    def api_stats():
        init_data = request.args.get("initData", "")
        user = _check_operator(init_data)
        if not user or user.get("id") not in ADMIN_IDS:
            return jsonify(ok=False), 403

        rows = get_all_operators_stats()
        items = [
            {
                "operator_name": r[0],
                "total": r[1],
                "done": r[2],
                "great": r[3] or 0,
                "ok": r[4] or 0,
                "bad": r[5] or 0,
            }
            for r in rows
        ]
        return jsonify(ok=True, items=items)

    # =========================
    # ЭКСПОРТ В EXCEL (только админ)
    # =========================
    @app.route("/api/dashboard/export")
    def api_export():
        init_data = request.args.get("initData", "")
        user = _check_operator(init_data)
        if not user or user.get("id") not in ADMIN_IDS:
            return jsonify(ok=False), 403

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        rows = get_all_requests()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заявки"

        headers = ["ID", "User ID", "Ресторан", "Описание", "Статус", "Оператор", "Оценка", "Причина"]
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value).alignment = Alignment(wrap_text=True)

        column_widths = [6, 14, 20, 40, 20, 15, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="zayavki.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
