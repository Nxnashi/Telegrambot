import logging
import io
from config import ADMIN_IDS

print(f"[DEBUG] ADMIN_IDS при загрузке модуля: {ADMIN_IDS}")

logger = logging.getLogger(__name__)


def register_admin_handlers(bot):

    print("[DEBUG] register_admin_handlers вызван")


    # =========================
    # /admin — главное меню
    # =========================
    @bot.message_handler(commands=["admin"])
    def admin_menu(message):
        print(f"[DEBUG] /admin от {message.from_user.id}, ADMIN_IDS: {ADMIN_IDS}")
        if message.from_user.id not in ADMIN_IDS:
            return

        from telebot.types import ReplyKeyboardMarkup, KeyboardButton
        markup = ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(KeyboardButton("📊 Статистика операторов"))
        markup.add(KeyboardButton("📥 Экспорт заявок в Excel"))
        markup.add(KeyboardButton("📋 Активные заявки"))
        markup.add(KeyboardButton("🗂 Доска заявок"))

        from telebot.types import ReplyKeyboardRemove
        bot.send_message(message.chat.id, "...", reply_markup=ReplyKeyboardRemove())
        bot.send_message(message.chat.id, "👑 Панель администратора", reply_markup=markup)
        print(f"[DEBUG] markup отправлен: {markup}")

    # =========================
    # Статистика всех операторов
    # =========================
    @bot.message_handler(func=lambda m: (m.text or "") == "📊 Статистика операторов" and m.from_user.id in ADMIN_IDS)
    def all_stats(message):
        from database import get_all_operators_stats
        rows = get_all_operators_stats()

        if not rows:
            bot.send_message(message.chat.id, "Нет данных по операторам.")
            return

        text = "📊 Статистика всех операторов:\n\n"

        for row in rows:
            operator_name, total, done, great, ok, bad = row
            in_progress = total - done
            text += f"👨‍💼 {operator_name}\n"
            text += f"  📋 Всего: {total}\n"
            text += f"  ✅ Выполнено: {done}\n"
            text += f"  🔄 В работе: {in_progress}\n"
            text += f"  ⭐ Отлично: {great or 0} | Нормально: {ok or 0} | Плохо: {bad or 0}\n"
            text += "\n"

        bot.send_message(message.chat.id, text)

    # =========================
    # Экспорт в Excel
    # =========================
    @bot.message_handler(func=lambda m: (m.text or "") == "📥 Экспорт заявок в Excel" and m.from_user.id in ADMIN_IDS)
    def export_excel(message):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from database import get_all_requests

            rows = get_all_requests()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Заявки"

            # Заголовки
            headers = ["ID", "User ID", "Ресторан", "Описание", "Статус", "Оператор", "Оценка"]
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Данные
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True)

                # Цвет строки по статусу
                status = row[4]
                if status == "Выполнено":
                    fill = PatternFill("solid", fgColor="E2EFDA")
                elif status == "Заявка в процессе":
                    fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    fill = PatternFill("solid", fgColor="FCE4D6")

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

            # Ширина колонок
            column_widths = [6, 14, 20, 40, 20, 15, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

            # Сохраняем в память и отправляем
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            bot.send_document(
                message.chat.id,
                ("zaявки.xlsx", buffer),
                caption=f"📥 Экспорт заявок — всего {len(rows)} записей"
            )

        except ImportError:
            bot.send_message(message.chat.id, "❌ Установи openpyxl: pip install openpyxl")
        except Exception as e:
            logger.error(f"Ошибка экспорта Excel: {e}")
            bot.send_message(message.chat.id, f"❌ Ошибка: {e}")

    # =========================
    # Активные заявки (для админа)
    # =========================
    @bot.message_handler(func=lambda m: (m.text or "") == "📋 Активные заявки" and m.from_user.id in ADMIN_IDS)
    def admin_active(message):
        from database import get_active_requests
        rows = get_active_requests()

        if not rows:
            bot.send_message(message.chat.id, "Нет активных заявок ✅")
            return

        text = "📋 Активные заявки:\n\n"
        for row in rows:
            request_id, restaurant, status, operator_name = row
            text += f"📌 Заявка #{request_id}\n"
            text += f"🏪 Ресторан: {restaurant}\n"
            text += f"📊 Статус: {status}\n"
            if operator_name:
                text += f"👨‍💼 Оператор: {operator_name}\n"
            text += "\n"

        bot.send_message(message.chat.id, text)
